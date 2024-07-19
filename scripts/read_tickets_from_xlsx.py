import openpyxl
from openpyxl import load_workbook


def read_seats_from_xlsx(name_document: str = "document.xlsx", read_sheet: str = 'Shemas', event_id: int = 1):
    data = []
    rows = []
    seats = []

    try:
        wb = load_workbook(name_document)
        work_sheet = wb[read_sheet]

        # Get the seat names from the first row, starting from the second column
        for cell in work_sheet[1][1:]:
            if cell.value is None:
                break
            seats.append(cell.value)

        # Get the row names and data from the sheet
        for row in work_sheet.iter_rows(min_row=2, values_only=True):
            row_name = row[0]
            if row_name is None:
                break
            rows.append(row_name)
            for seat_index, seat_value in enumerate(row[1:], start=1):
                if seat_index - 1 >= len(seats):
                    break
                if seat_value is not None:
                    data.append({
                        'event_id': event_id,
                        'row': row_name,
                        'seat': seats[seat_index - 1],
                        'category_id': seat_value
                    })

        # # Print row and seat names
        # print(f"Rows: {rows}")
        # print(f"Seats: {seats}")
        #
        # # Print the data in the specified format
        # for entry in data:
        #     print(entry)

    except Exception as e:
        print(e)

    return data


if __name__ == '__main__':
    result = read_seats_from_xlsx()
    print(result)
